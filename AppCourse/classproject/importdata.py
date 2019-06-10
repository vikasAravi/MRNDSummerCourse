import os
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'classproject.settings')
django.setup()



import click
import openpyxl
from onlineapp.models import *;

@click.group()
def hello():
    pass


@hello.command('collegedata')
@click.argument('input_arg', nargs = 1)
def collegeData(input_arg):
    sheet = openpyxl.load_workbook(input_arg)
    college_sheet = sheet.get_sheet_by_name('Colleges')
    row_count = college_sheet.max_row
    for i in range(2, row_count + 1):
        name = college_sheet.cell(row = i, column = 1).value
        acronym = college_sheet.cell(row = i, column = 2).value
        location = college_sheet.cell(row = i, column = 3).value
        contact = college_sheet.cell(row = i, column = 4).value
        c = College(name = name , acronym = acronym, location = location, contact = contact)
        c.save()


@hello.command('studentdata')
@click.argument('input', nargs = 1)
def studentData(input):
    sheet = openpyxl.load_workbook(input)
    student_sheet = sheet.get_sheet_by_name('Current')
    row_count = student_sheet.max_row
    for i in range(2, row_count + 1):
        name = student_sheet.cell(row = i, column = 1).value
        email = student_sheet.cell(row = i, column = 3).value
        dbnames = student_sheet.cell(row = i, column = 4).value
        college_id = College.objects.filter(acronym = student_sheet.cell(row = i,column = 2).value)[0]
        c = Student(name = name, email = email, db_folder = dbnames, college = college_id)
        c.save()


@hello.command('marksdata')
@click.argument('input', nargs = 1)
def marksdata(input):
    sheet = openpyxl.load_workbook(input)
    marks_sheet = sheet.get_sheet_by_name('Sheet')
    row_count = marks_sheet.max_row
    for i in range(2, row_count + 1):
        problem1 = marks_sheet.cell(row = i, column = 2).value
        problem2 = marks_sheet.cell(row = i, column = 3).value
        problem3 = marks_sheet.cell(row = i, column = 4).value
        problem4 = marks_sheet.cell(row = i, column = 5).value
        total =  marks_sheet.cell(row = i, column = 6).value
        marks_data = marks_sheet.cell(row=i, column=1).value
        clg_end_idx = marks_data[7:].index("_")
        college = marks_data[7:7 + clg_end_idx]
        name = marks_data[7 + clg_end_idx + 1:-5:].lower()
        student = (Student.objects.filter(db_folder = name))
        if(len(student)):
            c = MockTest1(student = student[0], problem1 = problem1, problem2= problem2, problem3 = problem3, problem4 = problem4, total = total)
            c.save()

if __name__ == '__main__':
    hello()